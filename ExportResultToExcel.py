# -*- coding: utf-8 -*-
import pandas as pd
from sqlalchemy import create_engine, text
import argparse
import configparser
import os
import sys
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('export_result.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def parse_args():
    parser = argparse.ArgumentParser(description="导出统计表到Excel")
    parser.add_argument('--host', help='数据库主机')
    parser.add_argument('--port', type=int, help='数据库端口')
    parser.add_argument('--dbname', help='数据库名')
    parser.add_argument('--user', help='数据库用户名')
    parser.add_argument('--password', help='数据库密码')
    return parser.parse_args()

def load_config():
    # 优先当前工作目录
    config_path = os.path.join(os.getcwd(), 'config.ini')
    if not os.path.exists(config_path):
        # 兼容未打包时
        config_path = os.path.join(os.path.dirname(sys.argv[0]), 'config.ini')
    if not os.path.exists(config_path):
        raise FileNotFoundError("config.ini 未找到")
    try:
        config = configparser.ConfigParser()
        config.read(config_path, encoding='utf-8')
        if 'database' not in config:
            logger.error("配置文件中缺少 [database] 部分")
            raise KeyError("缺少 [database] 部分")
        db = config['database']
        db_config = {
            'host': db.get('host', 'localhost'),
            'port': db.getint('port', 5432),
            'dbname': db.get('dbname'),
            'user': db.get('user'),
            'password': db.get('password')
        }
        missing = [key for key in ['dbname', 'user'] if not db_config[key]]
        if missing:
            logger.error("配置文件中缺少必要的数据库信息: %s", ", ".join(missing))
            raise ValueError("缺少必要的数据库配置")
        logger.info("已成功加载数据库配置")
        return db_config
    except Exception as e:
        logger.error("加载配置文件时出错: %s", str(e))
        raise

def get_db_config(args):
    if args.host and args.port and args.dbname and args.user and args.password:
        return {
            'host': args.host,
            'port': args.port,
            'dbname': args.dbname,
            'user': args.user,
            'password': args.password
        }
    return load_config()

def format_datetime_columns(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
    return df

def autofit_column_width(worksheet, df):
    for i, col in enumerate(df.columns, 1):
        max_length = max(
            df[col].astype(str).map(len).max() if not df[col].isnull().all() else 0,
            len(str(col))
        )
        worksheet.column_dimensions[get_column_letter(i)].width = max_length + 2

def set_worksheet_font(worksheet, df, font_name="等线", font_size=11):
    font = Font(name=font_name, size=font_size)
    # 设置表头
    for cell in worksheet[1]:
        cell.font = font
    # 设置数据区
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.font = font

if __name__ == "__main__":
    args = parse_args()
    try:
        db_config = get_db_config(args)
    except Exception as e:
        logger.error(f"数据库配置加载失败: {e}")
        sys.exit(1)

    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    stat_filename = f"统计结果_{now_str}.xlsx"
    details_filename = f"数据明细_{now_str}.xlsx"

    table_names = [
        'public.p_send_time',
        'public.p_epicenter_deviation',
        'public.p_judge_time',
        'public.p_mag_deviation',
        'public.p_warning_miss',
        'public.s_40gal_send_time',
        'public.s_80gal_send_time',
        'public.s_120gal_send_time',
        'public.s_warning_miss',
        'public.s_alarm_before_p',
        'public.s_40gal_judge_time',
        'public.s_80gal_judge_time',
        'public.s_120gal_judge_time',
        'public.s_peak_deviation'
    ]

    # 使用 SQLAlchemy 创建 engine
    engine = create_engine(
        f"postgresql+psycopg2://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}/{db_config['dbname']}"
    )

    try:
        with engine.connect() as conn:
            logger.info("正在连接数据库...")
            conn.execute(text("SELECT 1"))
            logger.info("数据库连接成功！")
    except Exception as e:
        logger.error(f"数据库连接失败: {e}")
        sys.exit(1)

    try:
        with pd.ExcelWriter(stat_filename, engine='openpyxl') as writer:
            for table in table_names:
                try:
                    df = pd.read_sql_query(f'SELECT * FROM {table}', engine)
                    df = format_datetime_columns(df)
                    if not df.empty:
                        sheet_name = str(df.iloc[0, 0])[:31]
                    else:
                        sheet_name = table.split('.')[-1]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]
                    autofit_column_width(ws, df)
                    set_worksheet_font(ws, df)  # 设置字体
                    logger.info(f"导出表 {table} 成功")
                except Exception as e:
                    logger.error(f"导出表 {table} 失败: {e}")

        logger.info("统计结果导出完成！")
    except Exception as e:
        logger.error(f"统计结果导出失败: {e}")
        sys.exit(1)
    
    # 导出 public.details 到单独的 Excel 文件
    try:
        df_details = pd.read_sql_query('SELECT * FROM public.details', engine)
        df_details = format_datetime_columns(df_details)
        with pd.ExcelWriter(details_filename, engine='openpyxl') as writer:
            df_details.to_excel(writer, sheet_name='details', index=False)
            ws = writer.sheets['details']
            autofit_column_width(ws, df_details)
            set_worksheet_font(ws, df_details)  # 设置字体
        logger.info("public.details 导出完成！")
    except Exception as e:
        logger.error(f"导出表 public.details 失败: {e}")
        sys.exit(1)